import openpyxl
import psycopg2
import getpass

def func1():
    file_location=input("Enter the file path of your excel sheet: ")
    workbook=openpyxl.load_workbook(file_location)
    w_sheet=input("Enter the name of worksheet: ")
    sheet=workbook[w_sheet]
    dbase=input("Enter the name of your database: ")
    user1=input("Enter the name of user: ")
    pword=getpass.getpass("Enter password: ")

    try:
        connection=psycopg2.connect(database=dbase,user=user1,password=pword,host="127.0.0.1",port="5433")
    except psycopg2.Error as err:
        print("An error was generated!!!")
    else:
        print("Connection to database was successful!")

    cursor=connection.cursor()

    schema =input("Enter the name of schema in which you want to make the table: ")
    table_name =input("Enter table name: ")

    cursor.execute('''create table {}.{}
                    (id int primary key not null,
                    first_name varchar(25) not null,
                    last_name varchar(25) not null,
                    department varchar(25) not null,
                    phone varchar(25),
                    address varchar(50),
                    salary int);'''.format(schema,table_name))

    query="""INSERT INTO {}.{}(id,first_name,last_name,department,phone,address,salary)VALUES(%s,%s,%s,%s,%s,%s,%s)""".format(schema,table_name)

    row1=int(input("Enter the first row number you want to migrate: "))
    row2=int(input("Enter the last row number you want to migrate: "))

    col=sheet.max_column

    for i in range(row1,row2):
        li=[]
        for j in range(1,col+1):
            li.append(sheet.cell(i,j).value)

        cursor.execute(query,li)

    connection.commit()
    cursor.close()
    connection.close()

    print("Migration complete!!!")

func1()
